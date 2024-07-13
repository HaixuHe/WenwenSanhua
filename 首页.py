import streamlit as st
import re
import pandas as pd
from openpyxl import load_workbook


def extract_numbers(text):
    # 首先匹配所有数字段
    all_numbers = re.findall(r'\d+', text)
    filtered = ''
    # 过滤掉前面有字母的数字
    filtered_numbers = []
    for number in all_numbers:
        if re.search(r'[A-Za-z]', text.split(number)[0][-1:]):
            continue
        if len(number) >= 8:
            if number[:4] != '2024' and number[:4] != '2023':
                filtered += str(number)
                filtered += '; '
                filtered_numbers.append(number)
    if len(filtered_numbers) == 1:
        filtered = filtered[:-2]
    return str(filtered)



st.set_page_config(layout="wide")

st.sidebar.info(
    """
    Copyright：杭州财务共享中心
    """
)

st.title("📷 发票号识别")

st.markdown(
    """- 上传指定文件，自动识别段落中发票号"""
)

point_path = st.file_uploader(
    "请选择需要加载Excel文件 👇",
    type=["XLSX", "XLS"],
)




print(point_path)
if point_path is not None:
    process_button = st.button('开始识别')
    if process_button:
        wb = load_workbook(filename=point_path, read_only=True)
        # print()
        ws = wb['工作表1']
        pattern = r'(?<![A-Za-z])\d{8,}'

        # 获取“摘要”列的所有数据
        VoucherNumberList = []
        SummaryList = []
        TaxList = []
        InvoiceNumberList = []
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            VoucherNumber, Summary, Tax = row[18], row[20], row[22]
            if VoucherNumber is not None and Summary is not None and Tax is not None and isinstance(Tax,
                                                                                                    (int, float,
                                                                                                     complex)) and (
                    '当前' not in Summary and VoucherNumber is not None):
                VoucherNumberList.append(VoucherNumber)
                SummaryList.append(Summary)
                TaxList.append(Tax)

                InvoiceNumberList.append(extract_numbers(Summary))
                # print(VoucherNumber, Summary, extract_numbers(Summary), Tax)

        df = pd.DataFrame(columns=['凭证号', '摘要', '发票号', '税额'])
        df['凭证号'] = VoucherNumberList
        df['摘要'] = SummaryList
        df['发票号'] = InvoiceNumberList
        df['税额'] = TaxList
        print(df)
        st.dataframe(df, use_container_width=True, hide_index=True)
        st.success('识别成功!', icon="✅")
        # df.to_excel('一般进项税-发票号导出.xlsx', index=False)
        # print(df)

